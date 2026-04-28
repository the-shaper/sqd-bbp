import json
import sys
from pathlib import Path


def limit_text(value: str, limit: int = 20000) -> str:
    cleaned = (value or "").strip()
    return cleaned[:limit]


def extract_pdf(file_path: Path) -> dict:
    from pypdf import PdfReader

    reader = PdfReader(str(file_path))
    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")

    text = limit_text("\n\n".join(pages))
    return {
        "extractedText": text,
        "summary": text[:500] if text else f"Uploaded PDF: {file_path.name}",
    }


def extract_docx(file_path: Path) -> dict:
    import docx

    document = docx.Document(str(file_path))
    paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
    text = limit_text("\n\n".join(paragraphs))
    return {
        "extractedText": text,
        "summary": text[:500] if text else f"Uploaded Word document: {file_path.name}",
    }


def extract_spreadsheet(file_path: Path) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(str(file_path), data_only=True)
    lines = []

    for sheet in workbook.worksheets[:5]:
        lines.append(f"[Sheet] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                lines.append(" | ".join(values))
            if len("\n".join(lines)) > 20000:
                break
        if len("\n".join(lines)) > 20000:
            break

    text = limit_text("\n".join(lines))
    return {
        "extractedText": text,
        "summary": text[:500] if text else f"Uploaded spreadsheet: {file_path.name}",
    }


def extract_text_file(file_path: Path) -> dict:
    text = limit_text(file_path.read_text(encoding="utf-8", errors="ignore"))
    return {
        "extractedText": text,
        "summary": text[:500] if text else f"Uploaded text document: {file_path.name}",
    }


def extract_image(file_path: Path) -> dict:
    from PIL import Image

    with Image.open(file_path) as image:
        width, height = image.size
        image_format = image.format or file_path.suffix.replace(".", "").upper() or "image"

    summary = (
        f"Uploaded image: {file_path.name} ({image_format}, {width}x{height}). "
        "Image OCR is not enabled yet, so this file is available as a visual reference."
    )
    return {
        "extractedText": "",
        "summary": summary,
    }


def main() -> None:
    file_path = Path(sys.argv[1])
    extension = file_path.suffix.lower()

    if extension == ".pdf":
        result = extract_pdf(file_path)
    elif extension in {".doc", ".docx"}:
        result = extract_docx(file_path)
    elif extension in {".xls", ".xlsx"}:
        result = extract_spreadsheet(file_path)
    elif extension in {".txt", ".md"}:
        result = extract_text_file(file_path)
    elif extension in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
        result = extract_image(file_path)
    else:
        result = {
            "extractedText": "",
            "summary": f"Uploaded file: {file_path.name}. This format is stored but not yet parsed.",
        }

    print(json.dumps(result))


if __name__ == "__main__":
    main()
